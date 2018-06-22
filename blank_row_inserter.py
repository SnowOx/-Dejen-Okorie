#! python3
# blank_row_inserter.py
# Terminal Usage Note:
#  -> blank_row_inserter.py <EXCEL_FILENAME> <row position> <number of blank rows to insert at the row position>

import openpyxl
import sys


EXCEL_FILENAME = sys.argv[1]


def get_workbook():
    workbook = openpyxl.load_workbook(EXCEL_FILENAME)
    return workbook


def get_sheetnames(workbook):
    sheetnames = workbook.sheetnames
    return sheetnames


def get_spreadsheet_contents(sheetname):
    spreadsheet_contents = []
    for column_number in range(1, sheetname.max_column + 1):
        inner_list = []
        for row_number in range(1, sheetname.max_row + 1):
            inner_list.append(sheetname.cell(row=row_number, column=column_number).value)
        spreadsheet_contents.append(inner_list)
    return spreadsheet_contents


def insert_blank_rows_into_spreadsheet_contents():
    for i in range(len(spreadsheet_contents)):
        for blank_column in range(int(sys.argv[3])):
            spreadsheet_contents[i].insert(int(sys.argv[2]), '')
    print(spreadsheet_contents)


def write_new_excel_document_based_on(sheetname, spreadsheet_contents):
    new_workbook = openpyxl.Workbook()
    new_workbook_sheet = new_workbook.active
    for i in range(len(spreadsheet_contents)):
        for j in range(len(spreadsheet_contents[i])):
            new_workbook_sheet.cell(row=j + 1, column=i + 1).value = spreadsheet_contents[i][j]
    save_new_workbook(new_workbook)


def save_new_workbook(new_workbook):
    new_workbook.save(
        str(EXCEL_FILENAME[:-5] + '_with_{}_blank_rows_inserted_at_row_{}.xlsx'.format(sys.argv[3], sys.argv[2])))


# Engine
workbook = get_workbook()
sheetnames = get_sheetnames(workbook)
for sheetname in workbook:
    spreadsheet_contents = get_spreadsheet_contents(sheetname)
    insert_blank_rows_into_spreadsheet_contents()
    write_new_excel_document_based_on(sheetname, spreadsheet_contents)

# Solved!
