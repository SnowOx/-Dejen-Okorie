#! python3
# excel_cell_inverter.py

import openpyxl
EXCEL_DOC = #

def get_workbook():
    workbook = openpyxl.load_workbook(EXCEL_DOC)
    return workbook

def get_sheetnames(workbook):
    sheetnames = workbook.sheetnames
    return sheetnames

def get_spreadsheet_contents(sheetname, workbook):
    spreadsheet_contents = []
    for column_number in range(1, sheetname.max_column + 1):
        inner_list = []
        for row_number in range(1, sheetname.max_row + 1):
            inner_list.append(sheetname.cell(
                row=row_number, column=column_number).value)
        spreadsheet_contents.append(inner_list)
    return spreadsheet_contents

def write_inverted_contents_based_on(sheetname, spreadsheet_contents):
    new_workbook = openpyxl.Workbook()
    new_workbook_sheet = new_workbook.active
    for i in range(len(spreadsheet_contents)):
        for j in range(len(spreadsheet_contents[i])):
            new_workbook_sheet.cell(row=i+1, column=j+1).value = spreadsheet_contents[i][j]
    save_new_workbook(new_workbook)
    
def save_new_workbook(new_workbook):
    new_workbook.save(str(xl_doc[:-5] + '_inverted.xlsx'))

# Engine
workbook = get_workbook()
sheetnames = get_sheetnames(workbook)
for sheetname in workbook:
    spreadsheet_contents = get_spreadsheet_contents(sheetname, workbook)
    write_inverted_contents_based_on(sheetname, spreadsheet_contents)
                         
# Solved
