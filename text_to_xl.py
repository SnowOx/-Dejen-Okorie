#! python3
# text_to_xl

import os
import openpyxl
from openpyxl.utils import get_column_letter
import random

FILE_FOLDER = 'C:\\Users\\A\\Google Drive\\OP2\\XLTestDocuments'
workbook = openpyxl.Workbook()
sheet = workbook['Sheet']

COLUMN_WIDTH = 80
ROW_HEIGHT = 26

def get_text_filenames_list(directory):
    text_filenames_list = [
        text_filename for text_filename in os.listdir(directory) 
        if text_filename.endswith('.txt')]
    return text_filenames_list

def put_text_content_into_list(text_filenames_list, directory):
    os.chdir(directory)
    text_contents_list = []
    for filename in text_filenames_list:
        read_content = open(filename).readlines()
        text_contents_list.append(read_content)
    return text_contents_list   

def find_length_of_longest_line(text_contents_list):
    counter = 0
    pprint.pprint(text_contents_list)
    for i in range(len(text_contents_list)):
        line_length = len(text_contents_list[i])
        if line_length > counter:
            counter = line_length
    length_of_longest_line = counter
    print(length_of_longest_line)
    return length_of_longest_line
    
def set_cell_dimensions(text_contents_list):
    for i in range(len(text_contents_list)):
        column_letter = get_column_letter(i+1)
        cell_column_dimensions = sheet.column_dimensions[column_letter]
        cell_column_dimensions.width = COLUMN_WIDTH
        for n in range(len(text_contents_list[i])):
            cell_row_dimensions = sheet.row_dimensions[n+1]
            cell_row_dimensions.height = ROW_HEIGHT  
   
def insert_text_and_save_xl_sheet(text_contents_list):
    for i in range(len(text_contents_list)):
        for n in range(len(text_contents_list[i])):
            cell_content = text_contents_list[i][n]
            cell_location = sheet.cell(column=i+1, row=n+1)
            cell_location.value = cell_content          
                
def save_workbook():
    workbook.save('testBook' + str(random.randint(0,1000)) + '.xlsx')

text_filenames_list = get_text_filenames_list(FILE_FOLDER)
text_contents_list = put_text_content_into_list(text_filenames_list, FILE_FOLDER)
set_cell_dimensions(text_contents_list)
insert_text_and_save_xl_sheet(text_contents_list)
save_workbook()

# Solved! I note that the row height and column widths are hardcoded.
# To automate column width formatting, I could find a ratio between the number of
# characters in each row and column width length and set the column widths based on this.
# To automate row height formatting, I could set the font for all text to insert into the xl sheet and set
# the row height based on this font.
