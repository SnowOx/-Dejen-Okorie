#! python3
# multiply_table

import sys
import openpyxl
from openpyxl.styles import Font

num = int(sys.argv[1])
workbook = openpyxl.Workbook()
sheet = workbook['Sheet']

def generate_headings():
    bold_font_object = Font(bold=True)
    for i in range(1, num + 1):
        col_heading = sheet.cell(row=1, column=i+1)
        row_heading = sheet.cell(row=i+1, column=1)
        col_heading.font = bold_font_object
        row_heading.font = bold_font_object # Use multiple assignment here to improve appearance?
        col_heading.value = i
        row_heading.value = i

def generate_table_content():
    for i in range(1, num + 1):
        for e in range(1, num + 1):
            cell_content = (sheet.cell(row=i+1, column=e).value) * (sheet.cell(row=i+1, column=e).value)
            cell_position = sheet.cell(row=i+1, column=e+1)
            cell_position.value = cell_content

def save_workbook():
    workbook.save('multiplication table for %s numbers.xlsx' % num)

generate_headings()
generate_table_content()
save_workbook()
