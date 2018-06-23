#! python3
# convert_spreadsheet_columns_to_text_files.py)

import openpyxl
import os


EXCEL_FILENAME = 'C:\\Users\\A\\Desktop\\test_document.xlsx'
SAVE_LOCATION = os.path.dirname(EXCEL_FILENAME)


def get_workbook():
    workbook = openpyxl.load_workbook(EXCEL_FILENAME)
    return workbook


def get_sheetnames(workbook):
    sheetnames = workbook.sheetnames
    return sheetnames


def get_spreadsheet_contents(sheetname):
    spreadsheet_contents = []
    for column_number in range(1, sheetname.max_column + 1):
        inner_list = []
        for row_number in range(1, sheetname.max_row + 1):
            inner_list.append(sheetname.cell(row=row_number, column=column_number).value)
        spreadsheet_contents.append(inner_list)
    return spreadsheet_contents


def write_document_based_on(sheetname, spreadsheet_contents):
    for counter,inner_list in enumerate(spreadsheet_contents):
        joined_inner_list = '\n'.join(str(cell) for cell in inner_list)
        print('Printing data from column %s to file' % counter)
        write_column_data_to_text_file(counter, joined_inner_list)
    print('Complete')


def write_column_data_to_text_file(counter, data_to_add_to_text_file):
    file_save_path = os.path.join(SAVE_LOCATION,'text_column' + str(counter) + '.txt')
    with open(file_save_path, 'w') as in_file:
        in_file.write('<<<< Column number %s data is below >>>\n' % str(counter))
        in_file.write(data_to_add_to_text_file)
        in_file.close()


# Engine
workbook = get_workbook()
sheetnames = get_sheetnames(workbook)
for sheetname in workbook:
    spreadsheet_contents = get_spreadsheet_contents(sheetname)
    write_document_based_on(sheetname, spreadsheet_contents)


#Solved :D