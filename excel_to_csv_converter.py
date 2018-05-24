#! python3
# excel_to_csv_converter.py

import openpyxl
import csv
import os

SEARCH_FOLDER = 'D:\\Test Folder'

def get_excel_files(location):
    excel_files = []
    for excel_file in os.listdir(location):
        if excel_file.endswith('.xlsx'):
            excel_files.append(excel_file)
    return excel_files

def get_excel_data(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    sheet_data = []
    for row_number in range(1, sheet.max_row):
        for column_number in range(1, sheet.max_column):
            sheet_data.append(
                sheet.cell(
                    row=row_number, column=column_number).value)
    return sheet_data

def write_sheet_data_to_csv(excel_file, sheet_data):
    output_file = open(str(excel_file[:-5]) + '.csv', 'w', newline='')
    output_writer = csv.writer(output_file)
    output_writer.writerow(sheet_data)
    output_file.close()

# Engine
os.chdir(SEARCH_FOLDER)
excel_files = get_excel_files(SEARCH_FOLDER)
for excel_file in excel_files:
    print(excel_file)
    sheet_data = get_excel_data(excel_file)
    print(sheet_data)
    write_sheet_data_to_csv(excel_file, sheet_data)
    print('Wrote %s to csv' % excel_file)

# Solved
